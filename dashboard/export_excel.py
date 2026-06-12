"""
导出 Dashboard 数据为 Excel（含原生图表）。

用法:
    python -m dashboard.export_excel
    python -m dashboard.export_excel -o data/cleaned/gaokao_report.xlsx
"""

from __future__ import annotations

import io
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import CLEANED_DIR
from dashboard.data_access import (
    get_category_breakdown,
    get_home_stats,
    get_quality_stats,
    get_school_export_data,
    get_top_schools,
    get_year_comparison,
)

HEADER_FONT = Font(bold=True)


def _autofit_columns(ws, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    if df.empty:
        ws.cell(row=start_row, column=1, value="(无数据)")
        return start_row + 1
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = HEADER_FONT
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    return start_row + len(df) + 1


def _add_grouped_bar_chart(
    ws,
    title: str,
    categories_col: int,
    data_start_col: int,
    data_end_col: int,
    data_start_row: int,
    data_end_row: int,
    anchor: str,
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = "数量"
    chart.x_axis.title = "年份"

    cats = Reference(
        ws,
        min_col=categories_col,
        min_row=data_start_row + 1,
        max_row=data_end_row,
    )
    data = Reference(
        ws,
        min_col=data_start_col,
        max_col=data_end_col,
        min_row=data_start_row,
        max_row=data_end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18
    ws.add_chart(chart, anchor)


def _add_top20_bar_chart(ws, data_start_row: int, data_end_row: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top20 高分院校（投档最低分）"
    chart.x_axis.title = "投档最低分"
    chart.y_axis.title = "院校"

    cats = Reference(
        ws,
        min_col=1,
        min_row=data_start_row + 1,
        max_row=data_end_row,
    )
    data = Reference(
        ws,
        min_col=2,
        min_row=data_start_row,
        max_row=data_end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 14
    chart.width = 16
    ws.add_chart(chart, anchor)


def _build_score_distribution(df: pd.DataFrame, bins: int = 20) -> pd.DataFrame:
    if df.empty or "min_score" not in df.columns:
        return pd.DataFrame(columns=["分数段", "院校数"])
    scores = pd.to_numeric(df["min_score"], errors="coerce").dropna()
    if scores.empty:
        return pd.DataFrame(columns=["分数段", "院校数"])
    hist, edges = pd.cut(scores, bins=bins, retbins=True)
    counts = hist.value_counts().sort_index()
    rows = []
    for interval, cnt in counts.items():
        left = int(interval.left)  # type: ignore[union-attr]
        right = int(interval.right)  # type: ignore[union-attr]
        rows.append({"分数段": f"{left}-{right}", "院校数": int(cnt)})
    return pd.DataFrame(rows)


def build_excel_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # --- 概览 ---
    ws_summary = wb.create_sheet("概览", 0)
    stats = get_home_stats()
    quality = get_quality_stats()
    ws_summary["A1"] = "高考录取线数据报告"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_df = pd.DataFrame(
        [
            {"指标": "school 记录数", "数值": stats["school_total"]},
            {"指标": "覆盖年份数", "数值": stats["year_count"]},
            {"指标": "覆盖省份数", "数值": stats["province_count"]},
            {"指标": "rank 记录数", "数值": quality["rank_total"]},
            {"指标": "control 记录数", "数值": quality["control_total"]},
        ]
    )
    _write_df(ws_summary, summary_df, start_row=4)
    _autofit_columns(ws_summary)

    # --- 年份对比 + 图表 ---
    ws_year = wb.create_sheet("年份对比")
    comparison = get_year_comparison([2023, 2024])
    if not comparison.empty:
        display = comparison.rename(
            columns={
                "year": "年份",
                "total_records": "总记录数",
                "history_count": "历史类数量",
                "physics_count": "物理类数量",
                "avg_min_score": "平均最低分",
                "max_min_score": "最高最低分",
            }
        )
        end_row = _write_df(ws_year, display)
        if len(display) >= 1:
            _add_grouped_bar_chart(
                ws_year,
                title="2023 vs 2024 科类记录数",
                categories_col=1,
                data_start_col=3,
                data_end_col=4,
                data_start_row=1,
                data_end_row=len(display) + 1,
                anchor="A10",
            )
    else:
        ws_year["A1"] = "暂无年份对比数据"
    _autofit_columns(ws_year)

    # --- 招生类别分布 ---
    ws_cat = wb.create_sheet("类别分布")
    cat_df = get_category_breakdown()
    cat_end = _write_df(ws_cat, cat_df)
    if not cat_df.empty and len(cat_df) <= 30:
        chart = BarChart()
        chart.type = "col"
        chart.title = "招生类别 × 批次 记录数"
        chart.y_axis.title = "记录数"
        cats = Reference(ws_cat, min_col=1, min_row=2, max_row=len(cat_df) + 1)
        data = Reference(ws_cat, min_col=3, min_row=1, max_row=len(cat_df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].data_labels = DataLabelList()
        chart.height = 12
        chart.width = 18
        ws_cat.add_chart(chart, f"A{cat_end + 2}")
    _autofit_columns(ws_cat)

    # --- 分数分布 + 图表 ---
    ws_dist = wb.create_sheet("分数分布")
    school_df = get_school_export_data()
    dist_df = _build_score_distribution(school_df)
    dist_end = _write_df(ws_dist, dist_df)
    if not dist_df.empty:
        chart = BarChart()
        chart.type = "col"
        chart.title = "投档最低分分布"
        chart.y_axis.title = "院校数"
        chart.x_axis.title = "分数段"
        cats = Reference(ws_dist, min_col=1, min_row=2, max_row=len(dist_df) + 1)
        data = Reference(ws_dist, min_col=2, min_row=1, max_row=len(dist_df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 20
        ws_dist.add_chart(chart, f"D2")
    _autofit_columns(ws_dist)

    # --- Top20 + 图表 ---
    ws_top = wb.create_sheet("Top20院校")
    top_df = get_top_schools(20)
    if not top_df.empty:
        top_display = top_df.rename(
            columns={
                "school_name": "院校",
                "min_score": "投档最低分",
                "subject_type": "科类",
                "year": "年份",
                "admission_category": "招生类别",
                "batch": "批次",
            }
        )
        top_display = top_display[
            ["院校", "投档最低分", "科类", "年份", "招生类别", "批次"]
        ]
        _write_df(ws_top, top_display)
        _add_top20_bar_chart(ws_top, 1, min(21, len(top_display) + 1), "F2")
    else:
        ws_top["A1"] = "暂无 Top20 数据"
    _autofit_columns(ws_top)

    # --- 全量院校数据 ---
    ws_data = wb.create_sheet("院校数据")
    export_df = school_df.rename(
        columns={
            "year": "年份",
            "province": "省份",
            "subject_type": "科类",
            "admission_category": "招生类别",
            "batch": "批次",
            "school_name": "院校名称",
            "school_code": "院校代号",
            "major_group": "专业组",
            "min_score": "投档最低分",
            "min_rank": "位次",
            "plan_count": "计划数",
        }
    )
    _write_df(ws_data, export_df)
    _autofit_columns(ws_data, max_width=50)

    return wb


def export_excel_bytes() -> bytes:
    buffer = io.BytesIO()
    wb = build_excel_workbook()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_excel_file(output_path: str | Path | None = None) -> Path:
    CLEANED_DIR.mkdir(parents=True, exist_ok=True)
    out = Path(output_path) if output_path else CLEANED_DIR / "gaokao_dashboard_report.xlsx"
    wb = build_excel_workbook()
    wb.save(out)
    return out


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="导出 Dashboard Excel 报告（含图表）")
    parser.add_argument("-o", "--output", help="输出路径")
    args = parser.parse_args()
    path = export_excel_file(args.output)
    print(f"已导出: {path}")
